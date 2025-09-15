"""
==================================================
Script: app.py
Author: Priyanka Gulgari
Date: 06-18-2024
Description: This script is a Flask web application designed to process pharmacy data. It allows users to upload multiple Excel files, including BestRx data, vendor data, and conversion data.
The script processes these files to aggregate and merge data, calculate package sizes, and generate a comprehensive report in Excel format.
The report includes a summary of purchased quantities,billed quantities, package size differences, and highlights any missing items that need to be updated in the master file.
The application supports optional insurance files and ensures data integrity throughout the process.
==================================================

License:
This script is the intellectual property of Priyanka Gulgari.
Unauthorized copying, distribution, modification, or use of this code, via any medium, is strictly prohibited without prior written permission from the author.

Contact:
For permissions or inquiries, please contact priyankagulgari@gmail.com .

==================================================
"""
import sys
from flask import Flask, request, redirect, url_for, send_file, render_template
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import csv
import webview
from flaskwebgui import FlaskUI
import tkinter as tk
from tkinter import filedialog
#import win32com.client as win32
from openpyxl.styles import numbers


root = tk.Tk()
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
root.destroy()

def resource_path(relative_path):
    """ Get the absolute path to the resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

app = Flask(__name__, template_folder='templates')
app.static_folder = 'static'
app_display_name = "app$"  # Use this for display purposes


#window = webview.create_window('Pharmacy Data Processing Application',app)
UPLOAD_FOLDER = 'uploads'
PROCESSED_FOLDER = 'processed'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
if not os.path.exists(PROCESSED_FOLDER):
    os.makedirs(PROCESSED_FOLDER)
    
    
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    
    pharmacy_name = request.form['pharmacy_name']
    date_range = request.form['date_range']
    insurance_files = {
        'ALL_PBM': request.files.get('bestrx_file'),
        'CVS': request.files.get('cvs_bestrx_file'),
        'ESI': request.files.get('esi_bestrx_file'),
        'OPTUM': request.files.get('optum_bestrx_file'),
        'MEDIMP': request.files.get('medimpact_bestrx_file'),
        'NYM': request.files.get('nym_bestrx_file'),
    }

    # Read optional insurances
    optional_insurance_count = int(request.form['optional_insurance_count'])
    for i in range(1, optional_insurance_count + 1):
        name = request.form[f'optional_insurance_name{i}']
        file = request.files.get(f'optional_insurance_file{i}')
        insurance_files[name] = file
        
    # Debugging print statement to verify optional insurances
    print("Insurance files:")
    for key, file in insurance_files.items():
        if file:
            print(f"{key}: {file.filename}")

            
    kinray_file = request.files['kinray_file']
    vendor_count = int(request.form['vendor_count'])                                  
    conversion_file = request.files['conversion_file']
                                    
    #Vendor files, dpending on user input
    vendor_files = []
    for i in range(1, vendor_count +1):
        vendor_name = request.form.get(f'vendor{i}_name', f'vendor{i}').strip()
        vendor_file = request.files.get(f'vendor{i}_file')     # from file input
        #vendor_data.append((vendor_name, vendor_file))
        #vendor_files.append(request.files[f'vendor{i}_file'])
        if vendor_file:
            vendor_files.append((vendor_name, vendor_file))
        
    if kinray_file.filename == '' or conversion_file.filename == '' or not any(insurance_files.values()):
        return redirect(request.url)
    
    insurance_paths = {}
    for key, file in insurance_files.items():
        if file:
            path = os.path.join(app.config['UPLOAD_FOLDER'], f'{key}.xlsx')
            file.save(path)
            insurance_paths[key] = path
            
    kinray_path = os.path.join(app.config['UPLOAD_FOLDER'], 'kinray.xlsx')
    conversion_path = os.path.join(app.config['UPLOAD_FOLDER'], 'conversion.xlsx')
    kinray_file.save(kinray_path)
    conversion_file.save(conversion_path)

    vendor_paths = []
    for i,(vendor_name, vendor_file) in enumerate(vendor_files, start=1):
        #if not vendor_name.strip():  # fallback if vendor name not entered
            #vendor_name = f'vendor{i}'
        safe_name = vendor_name.replace(" ", "_") or f'vendor{i}'  # fallback if empty
        vendor_path = os.path.join(app.config['UPLOAD_FOLDER'], f'{safe_name}.xlsx')
        vendor_file.save(vendor_path)
        vendor_paths.append(vendor_path)
        #vendor_path = os.path.join(app.config['UPLOAD_FOLDER'], f'vendor{i}.xlsx')
        #file.save(vendor_path)
        #vendor_paths.append(vendor_path)

    processed_file_path= process_files(insurance_paths, [kinray_path] + vendor_paths, conversion_path, pharmacy_name, date_range)
        
    # Debugging print statements
    # Ensure the file exists before sending it
    if os.path.exists(processed_file_path):
        send_file(processed_file_path, as_attachment=True)
        return render_template('success.html', message="Your file has been downloaded successfully.")
    else:
        return "Error: File not found."
        
    return send_file(processed_file_path, as_attachment=True)


def add_missing_items_sheet(wb, missing_items):
    ws_missing = wb.create_sheet(title="Missing Items")

    # Set the header
    ws_missing.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(missing_items.columns))
    cell = ws_missing.cell(row=1, column=1)
    cell.value = "Missing items, To be updated in master file"
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(size=20, bold=True)
    ws_missing.row_dimensions[1].height = 30
    # Add the missing items data
    for r_idx, row in enumerate(dataframe_to_rows(missing_items, index=False, header=True), start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws_missing.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=12)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    # Set fixed column widths for specific columns
    column_widths = {
        'A': 20,  # Column A
        'B': 80,  # Column B
    }
    ws_missing.freeze_panes = 'A3'
    for col_letter, width in column_widths.items():
        ws_missing.column_dimensions[col_letter].width = width

    # Set the same width for other columns if needed
    for col in ws_missing.columns:
        column = col[0].coordinate[:1]  # Get the column letter
        if column not in column_widths:  # If not already set, apply a default width
            ws_missing.column_dimensions[column].width = 20
def add_needs_to_order_sheet(wb, final_data, conversion_data):
    #Filter for rows where CVS_D (CVS Package Size Difference) is negative
    #price_mapping = conversion_data.set_index('NDC #')['PRICE'].to_dict()
    #final_data['PRICE'] = final_data['NDC #'].map(price_mapping)

    needs_to_order = final_data[final_data['CVS_D'] < 0][['NDC #', 'Drug Name','Package Size', 'CVS_D', 'PRICE', 'Total Order Price']].copy()

    #needs_to_order['Total Order Price'] = needs_to_order['PRICE'] * needs_to_order['CVS_D']
    
    
    if needs_to_order.empty:
        return  # If no negative values, return without adding the sheet
    needs_to_order = needs_to_order.sort_values(by='Drug Name')

    needs_to_order.rename(columns={'Package Size': 'Pkg Size'}, inplace=True)

    needs_to_order.insert(needs_to_order.columns.get_loc('CVS_D') + 1, 'Paper Work', '')
    

    # Create a new sheet for needs to be ordered
    ws_needs_order = wb.create_sheet(title="Needs to be Ordered CVS")

    # Set the header
    ws_needs_order.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(needs_to_order.columns))
    cell = ws_needs_order.cell(row=1, column=1)
    cell.value = "Needs to be ordered CVS"
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(size=20, bold=True)
    ws_needs_order.row_dimensions[1].height = 30

    # Add the data to the sheet
    for r_idx, row in enumerate(dataframe_to_rows(needs_to_order, index=False, header=True), start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws_needs_order.cell(row=r_idx, column=c_idx, value=value)
            if c_idx in [2,5,6] :
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:   
                cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=12)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Set column widths for the new sheet
    column_widths = {
        'A': 15,  # NDC #
        'B': 70,  # Drug Name
        'C': 10,  # pkg size
        'D' : 15, # CVS_D (Difference)
        'E' : 20,
        'F' : 10,
        'G' : 16,
    }

    ws_needs_order.freeze_panes = 'A3'
    for col_letter, width in column_widths.items():
        ws_needs_order.column_dimensions[col_letter].width = width

    ws_needs_order.page_margins = PageMargins(left=0, right=0, top=0, bottom=0, header=0, footer=0)
    ws_needs_order.sheet_view.showGridLines = True
    ws_needs_order.print_options.gridLines = True
    ws_needs_order.print_options.horizontalCentered = False  # Do not center horizontally
    ws_needs_order.print_options.verticalCentered = False 

    ws_needs_order.page_setup.fitToHeight = False  # Ensure it fits all rows vertically
    ws_needs_order.page_setup.fitToWidth = 1

    #ws_needs_order.page_setup.orientation = ws_needs_order.ORIENTATION_LANDSCAPE

    ws_needs_order.print_options.gridLines = True

    total_rows = ws_needs_order.max_row + 1  # Start after the last row of data
    ws_needs_order.cell(row=total_rows, column=6).value = "Total Order Price"
    ws_needs_order.cell(row=total_rows, column=7).value = f"=SUM(G3:G{total_rows-1})"  # Formula for sum
     # Style the total sum row
    total_label_cell = ws_needs_order.cell(row=total_rows, column=6)
    total_value_cell = ws_needs_order.cell(row=total_rows, column=7)
    
    total_label_cell.font = Font(size=12, bold=True)
    total_value_cell.font = Font(size=12, bold=True)
    total_value_cell.number_format = '"$"#,##0.00'  # Format as currency

    total_label_cell.alignment = Alignment(horizontal='center')
    total_value_cell.alignment = Alignment(horizontal='center')

def add_do_not_order(wb, final_data):
    # Filter for rows where CVS_D (CVS Package Size Difference) is negative
    needs_to_order = final_data[final_data['CVS_D'] > 0][['NDC #', 'Drug Name','Package Size', 'CVS_D']].copy()
    if needs_to_order.empty:
        return  # If no negative values, return without adding the sheet
    needs_to_order = needs_to_order.sort_values(by='Drug Name')

    needs_to_order.rename(columns={'Package Size': 'Pkg Size'}, inplace=True)

    needs_to_order['Paper Work'] = ''  # Add an empty column for "Paper Work"


    # Create a new sheet for needs to be ordered
    ws_needs_order = wb.create_sheet(title="Do Not Order CVS")

    # Set the header
    ws_needs_order.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(needs_to_order.columns))
    cell = ws_needs_order.cell(row=1, column=1)
    cell.value = "Do Not Order CVS"
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(size=20, bold=True)
    ws_needs_order.row_dimensions[1].height = 30

    # Add the data to the sheet
    for r_idx, row in enumerate(dataframe_to_rows(needs_to_order, index=False, header=True), start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws_needs_order.cell(row=r_idx, column=c_idx, value=value)
            if c_idx == 2:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:   
                cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=12)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Set column widths for the new sheet
    column_widths = {
        'A': 15,  # NDC #
        'B': 70,  # Drug Name
        'C': 10,  # pkg size
        'D' : 15, # CVS_D (Difference)
        'E' : 20, # Paper work
    }

    ws_needs_order.freeze_panes = 'A3'
    for col_letter, width in column_widths.items():
        ws_needs_order.column_dimensions[col_letter].width = width

    ws_needs_order.page_margins = PageMargins(left=0, right=0, top=0, bottom=0, header=0, footer=0)
    ws_needs_order.print_options.horizontalCentered = False  # Do not center horizontally
    ws_needs_order.print_options.verticalCentered = False
    
    ws_needs_order.sheet_view.showGridLines = True
    ws_needs_order.print_options.gridLines = True

    ws_needs_order.page_setup.fitToHeight = False  # Ensure it fits all rows vertically
    ws_needs_order.page_setup.fitToWidth = 1

    #ws_needs_order.page_setup.orientation = ws_needs_order.ORIENTATION_LANDSCAPE
    ws_needs_order.print_options.gridLines = True
    
def get_column_index(ws, header_name):
    """
    Get the column index for a specific header name in the given worksheet.
    Args:
        ws: The worksheet object.
        header_name: The header name to search for.
    Returns:
        The column index (1-based) if found, otherwise None.
    """
    for cell in ws[3]:  # Assuming headers are in row 3
        if cell.value == header_name:
            return cell.column  # Return 1-based index
    return None

def add_autosum(ws, insurance_paths, start_row, end_row):
    # Loop through each insurance to add autosum for columns _T, _Pur, and _Diff$
    for insurance in insurance_paths.keys():
        # Get column indices dynamically for _T, _Pur, and _Diff$
        t_col = get_column_index(ws, f'{insurance}_T')
        pur_col = get_column_index(ws, f'{insurance}_Pur')
        diff_col = get_column_index(ws, f'{insurance}_Diff$')

        # Add autosum formula for each column
        if t_col:
            cell = ws.cell(row=end_row + 1, column=t_col)
            cell.value = f"=SUM({get_column_letter(t_col)}{start_row}:{get_column_letter(t_col)}{end_row})"
            cell.number_format = "#,##0"  # Format large numbers with commas
            cell.font = Font(size=12, bold=False)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        if pur_col:
            cell = ws.cell(row=end_row + 1, column=pur_col)
            cell.value = f"=SUM({get_column_letter(pur_col)}{start_row}:{get_column_letter(pur_col)}{end_row})"
            cell.number_format = "#,##0.00"  # Format as currency or decimal
            cell.font = Font(size=12, bold=False)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        if diff_col:
            cell = ws.cell(row=end_row + 1, column=diff_col)
            cell.value = f"=SUM({get_column_letter(diff_col)}{start_row}:{get_column_letter(diff_col)}{end_row})"
            cell.number_format = '"$"#,##0.00'  # Format as currency
            cell.font = Font(size=12, bold=False)
            cell.alignment = Alignment(horizontal='center', vertical='center')
                
def adjust_specific_columns(ws, columns_to_adjust):
    """
    Adjust the width of specific columns only.
    :param ws: Worksheet
    :param columns_to_adjust: List of column letters or indices to adjust
    """
    for col_letter in columns_to_adjust:
        max_length = 0
        for cell in ws[col_letter]:
            try:
                # Check the length of the value in the cell (convert to string)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length  # Exact fit, no padding
        
def add_max_difference_sheet(wb, final_data, insurance_paths):
    # Identify all difference columns ending with '_D'
    difference_columns = [col for col in final_data.columns if col.endswith('_D') and col != 'ALL_PBM_D']

    filtered_data = final_data.copy()
    filtered_data[difference_columns] = filtered_data[difference_columns].fillna(0)

    # Transform difference columns: Replace positive values with 0
    for col in difference_columns:
            filtered_data[col] = filtered_data[col].apply(lambda x: x if x < 0 else 0)

    # Filter rows where any of the selected `_D` columns have values less than 0
    needs_to_order = filtered_data[filtered_data[difference_columns].lt(0).any(axis=1)][['NDC #', 'Drug Name', 'Package Size'] + difference_columns + ['PRICE']].copy()

    # Create a new sheet for maximum differences
    ws_max_diff = wb.create_sheet(title="Needs to be ordered - All")
    
    if needs_to_order.empty:
        print("No rows with negative values in the selected difference columns.")
        return  # Return if no rows meet the condition

    # Ignore the negative sign and calculate the maximum absolute difference for each row
    needs_to_order['To Order'] = needs_to_order[difference_columns].abs().max(axis=1)
    needs_to_order['Paper Work'] = " "
    # Calculate the total order price (Max Difference * PRICE)
    needs_to_order['Total Order Price'] = needs_to_order['To Order'] * needs_to_order['PRICE']
    needs_to_order.rename(columns={'Package Size': 'Pkg Size'}, inplace=True)
    # Select the columns to display
    display_columns = ['NDC #', 'Drug Name', 'Pkg Size'] + difference_columns + ['To Order','Paper Work','PRICE', 'Total Order Price']
    
    # Sort by Drug Name for better readability
    needs_to_order = needs_to_order[display_columns].sort_values(by='Drug Name')
        
    # Set the header
    ws_max_diff.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(display_columns))
    cell = ws_max_diff.cell(row=1, column=1)
    cell.value = "Differences Across Insurances"
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(size=20, bold=True)
    ws_max_diff.row_dimensions[1].height = 30
    
    # Add the data to the sheet
    for r_idx, row in enumerate(dataframe_to_rows(needs_to_order, index=False, header=True), start=2):
        for c_idx, value in enumerate(row, start=1):
            
            cell = ws_max_diff.cell(row=r_idx, column=c_idx, value=value)
            
            if c_idx == display_columns.index("Drug Name") + 1:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')

            cell.font = Font(size=12)

    # Set column width for "Paper Work"
    paper_work_col_idx = display_columns.index("Paper Work") + 1
    ws_max_diff.column_dimensions[get_column_letter(paper_work_col_idx)].width = 10  # Adjust as needed


    # Apply thick borders to Row 2 (Header Row)
    for col_idx in range(1, len(display_columns) + 1):
        cell = ws_max_diff.cell(row=2, column=col_idx)
        cell.border = Border(
            top=Side(style='thick'),
            left=Side(style='thick'),
            right=Side(style='thick'),
            bottom=Side(style='thick')
        )

    # Apply thick borders to column edges only
    def apply_column_border(ws, col_idx):
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
            cell = ws[f"{col_letter}{row[0].row}"]
            cell.border = Border(left=Side(style='thick'), right=Side(style='thick'))

    # Apply thick borders to specific columns
    thick_border_columns = ['NDC #', 'Drug Name', 'Pkg Size', 'PRICE', 'To Order', 'Total Order Price','Paper Work']
    for col_name in thick_border_columns:
        if col_name in display_columns:
            apply_column_border(ws_max_diff, display_columns.index(col_name) + 1)

    # Set column widths dynamically
    insurance_columns = difference_columns
    for col_name in insurance_columns:
        if col_name in display_columns:
            col_idx = display_columns.index(col_name) + 1
            col_letter = get_column_letter(col_idx)
            ws_max_diff.column_dimensions[col_letter].width = 8  # Insurance columns
            header_cell = ws_max_diff.cell(row=2, column=col_idx)
            header_cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)

        else:
            print(f"Warning: Column {col_name} not found in display_columns.")

    column_widths = {
        'A': 15,  # NDC #
        'B': 70,  # Drug Name
        'C': 7,  # Package Size
    }
    
    #ws.row_dimensions[2].height = 80
    for col_letter, width in column_widths.items():
        ws_max_diff.column_dimensions[col_letter].width = width
        
    ws_max_diff.freeze_panes = 'A3'
    ws_max_diff.page_setup.orientation = ws_max_diff.ORIENTATION_PORTRAIT

    to_order_col_idx = display_columns.index("To Order") + 1
    to_order_col_letter = get_column_letter(to_order_col_idx)
    
    ws_max_diff.column_dimensions[to_order_col_letter].width = 15  # "To Order" column
    ws_max_diff.column_dimensions[get_column_letter(len(display_columns) - 1)].width = 15  # PRICE
    ws_max_diff.column_dimensions[get_column_letter(len(display_columns))].width = 20  # Total Order Price

    # Adjust row height for row 2
    ws_max_diff.row_dimensions[2].height = 80


    # Add a total row for Total Order Price
    total_rows = ws_max_diff.max_row + 1  # Start after the last row of data
    ws_max_diff.cell(row=total_rows, column=len(display_columns) - 1).value = "Total Order Price"
    ws_max_diff.cell(row=total_rows, column=len(display_columns)).value = f"=SUM({get_column_letter(len(display_columns))}2:{get_column_letter(len(display_columns))}{total_rows-1})"

    # Style the total row
    total_label_cell = ws_max_diff.cell(row=total_rows, column=len(display_columns) - 1)
    total_value_cell = ws_max_diff.cell(row=total_rows, column=len(display_columns))

    total_label_cell.font = Font(size=12, bold=True)
    total_value_cell.font = Font(size=12, bold=True)
    total_value_cell.number_format = '"$"#,##0.00'  # Format as currency

    total_label_cell.alignment = Alignment(horizontal='center', vertical='center')
    total_value_cell.alignment = Alignment(horizontal='center', vertical='center')

def min_difference_sheet(wb, final_data, insurance_paths):
    # Identify all difference columns ending with '_D'
    difference_columns = [col for col in final_data.columns if col.endswith('_D') and col != 'ALL_PBM_D']

    filtered_data = final_data.copy()
    filtered_data[difference_columns] = filtered_data[difference_columns].fillna(0)


    # Define a threshold for rounding small values to zero
    #epsilon = 1e-4  # You can adjust this value as needed
    #filtered_data[difference_columns] = filtered_data[difference_columns].applymap(lambda x: 0 if abs(x) < epsilon else x)
    # Transform difference columns: Replace negative values with 0
    for col in difference_columns:
            filtered_data[col] = filtered_data[col].apply(lambda x: x if x > 0 else 0)

    filtered_data['Min Positive'] = filtered_data[difference_columns].min(axis=1)

    # Filter rows where any of the selected `_D` columns have values less than 0
    do_not_order = filtered_data[filtered_data['Min Positive'] > 0][['NDC #', 'Drug Name', 'Package Size'] + difference_columns + ['Min Positive', 'PRICE']].copy()

    #do_not_order = filtered_data[['NDC #', 'Drug Name', 'Package Size'] + difference_columns + ['Min Positive', 'PRICE']].copy()

    # Create a new sheet for maximum differences
    ws_max_diff = wb.create_sheet(title="Do Not Order - ALL")#Do Not Order - All
    
    if do_not_order.empty:
        print("No rows with negative values in the selected difference columns.")
        return  # Return if no rows meet the condition

    do_not_order['Paper\nWork'] = " "
    do_not_order.rename(columns={'Package Size': 'Pkg Size'}, inplace=True)
    # Select the columns to display
    display_columns = ['NDC #', 'Drug Name', 'Pkg Size'] + difference_columns + ['Min Positive', 'Paper\nWork']
    
    # Sort by Drug Name for better readability
    do_not_order = do_not_order[display_columns].sort_values(by='Drug Name')
        
    # Set the header
    ws_max_diff.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(display_columns))
    cell = ws_max_diff.cell(row=1, column=1)
    cell.value = "Do not order"
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(size=20, bold=True)
    ws_max_diff.row_dimensions[1].height = 30
    
    # Add the data to the sheet
    for r_idx, row in enumerate(dataframe_to_rows(do_not_order, index=False, header=True), start=2):
        for c_idx, value in enumerate(row, start=1):
            
            cell = ws_max_diff.cell(row=r_idx, column=c_idx, value=value)
            if c_idx == display_columns.index("Drug Name") + 1:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')

            cell.font = Font(size=12)

    # Set column width for "Paper Work"
    paper_work_col_idx = display_columns.index("Paper\nWork") + 1
    ws_max_diff.column_dimensions[get_column_letter(paper_work_col_idx)].width = 10  # Adjust as needed

    # Apply thick borders to Row 2 (Header Row)
    for col_idx in range(1, len(display_columns) + 1):
        cell = ws_max_diff.cell(row=2, column=col_idx)
        cell.border = Border(
            top=Side(style='thick'),
            left=Side(style='thick'),
            right=Side(style='thick'),
            bottom=Side(style='thick')
        )

    # Apply thick borders to column edges only
    def apply_column_border(ws, col_idx):
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
            cell = ws[f"{col_letter}{row[0].row}"]
            cell.border = Border(left=Side(style='thick'), right=Side(style='thick'))

    # Apply thick borders to specific columns
    thick_border_columns = ['NDC #', 'Drug Name', 'Pkg Size','Min Positive','Paper\nWork']
    for col_name in thick_border_columns:
        if col_name in display_columns:
            apply_column_border(ws_max_diff, display_columns.index(col_name) + 1)

    # Set column widths dynamically
    insurance_columns = difference_columns
    for col_name in insurance_columns:
        if col_name in display_columns:
            col_idx = display_columns.index(col_name) + 1
            col_letter = get_column_letter(col_idx)
            ws_max_diff.column_dimensions[col_letter].width = 8  # Insurance columns
            header_cell = ws_max_diff.cell(row=2, column=col_idx)
            header_cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)

        else:
            print(f"Warning: Column {col_name} not found in display_columns.")

    rotated_columns = ['Pkg Size', 'Min Positive']
    for col_name in rotated_columns:
        if col_name in display_columns:
            col_idx = display_columns.index(col_name) + 1
            col_letter = get_column_letter(col_idx)
            ws_max_diff.column_dimensions[col_letter].width = 8
            header_cell = ws_max_diff.cell(row=2, column=col_idx)
            header_cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)


    column_widths = {
        'A': 15,  # NDC #
        'B': 60,  # Drug Name
        'C': 7,  # Package Size
    }
    
    #ws.row_dimensions[2].height = 80
    for col_letter, width in column_widths.items():
        ws_max_diff.column_dimensions[col_letter].width = width
        
    ws_max_diff.freeze_panes = 'A3'

    # Adjust row height for row 2
    ws_max_diff.row_dimensions[2].height = 80

# Main Function to Create "Never Ordered - Check" Sheet
def create_never_ordered_check_sheet(wb, final_data):
    # Filter rows where 'Total Purchased' is 0
    never_ordered_data = final_data[final_data['Total Purchased'] == 0]

    # Select the required columns and include quantity billed columns for all insurance paths
    insurance_columns = [col for col in final_data.columns if col.endswith('_P')]
    columns_to_select = ['Drug Name', 'NDC #', 'Package Size', 'Total Purchased'] + insurance_columns
    never_ordered_data = never_ordered_data[columns_to_select]
    never_ordered_data.rename(columns={'Package Size': 'Pkg Size'}, inplace=True)

    if never_ordered_data.empty:
        print("No rows with Total Purchased = 0 to report.")
        return

    never_ordered_data = never_ordered_data.sort_values(by='Drug Name')

    # Create a new sheet in the workbook
    ws = wb.create_sheet(title="Never Ordered - Check")

    # Set the header
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(never_ordered_data.columns))
    cell = ws.cell(row=1, column=1)
    cell.value = "Never Ordered - Check"
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(size=20, bold=True)
    ws.row_dimensions[1].height = 30

    # Add the data to the sheet
    for r_idx, row in enumerate(dataframe_to_rows(never_ordered_data, index=False, header=True), start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if c_idx == 1:  # Drug Name column
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=12)

    # Apply thick borders to Row 2 (Header Row)
    for col_idx in range(1, len(columns_to_select) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.border = Border(
            top=Side(style='thick'),
            left=Side(style='thick'),
            right=Side(style='thick'),
            bottom=Side(style='thick')
        )

    # Apply thick borders to specific columns
    def apply_column_border(ws, col_idx):
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
            cell = ws[f"{col_letter}{row[0].row}"]
            cell.border = Border(left=Side(style='thick'), right=Side(style='thick'))

    thick_border_columns = ['Drug Name', 'NDC #', 'Pkg Size', 'Total Purchased']
    for col_name in thick_border_columns:
        if col_name in never_ordered_data.columns:
            col_idx = never_ordered_data.columns.get_loc(col_name) + 1
            apply_column_border(ws, col_idx)

    # Adjust column widths and header formatting
    column_widths = {'A': 70, 'B': 15, 'C': 10, 'D': 10}
    
    for idx, col_name in enumerate(never_ordered_data.columns, start=1):
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = column_widths.get(col_letter, 8)
        header_cell = ws.cell(row=2, column=idx)
        header_cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)

    ws.freeze_panes = 'A3'
    #print("Sheet 'Never Ordered - Check' created successfully.")

       
def process_files(insurance_paths, vendor_paths, conversion_path, pharmacy_name, date_range):
    
    dropped_data = []
    # Read data from BestRx software with NDC as string and necessary columns
    all_bestrx_data = []
    for insurance, path in insurance_paths.items():
        data = pd.read_excel(path, usecols=['Drug Name', 'NDC #', 'Total Rxs', 'Quantity','Total'], dtype={'NDC': str})
        print(f"Columns in {path}: {data.columns.tolist()}")
        data['Insurance'] = insurance
        all_bestrx_data.append(data)

        if 'Total' in data.columns:
            data['Total'] = data['Total'].round(0)
    #combined_bestrx_data = pd.concat(all_bestrx_data)
    combined_bestrx_data = pd.concat(all_bestrx_data)
    #print("Combined BestRx Data:")
    #print(combined_bestrx_data.head())
    #print("Combined BestRx Data shape:", combined_bestrx_data.shape)

    # Read data from Kinray vendor with NDC as string and necessary columns
    #kinray_data = pd.read_excel(kinray_path, usecols=['NDC', 'Shipped'], dtype={'NDC': str})
    
    
    # Read data from vendor files with NDC as string and necessary columns
    all_vendor_data = []
    vendor_names=[]
    for vendor_index, vendor_path in enumerate(vendor_paths, start =1):
        vendor_data = pd.read_excel(vendor_path, usecols=['NDC #', 'Shipped'], dtype={'NDC #': str})
        vendor_data['Vendor'] = f'Vendor{vendor_index}'
        all_vendor_data.append(vendor_data)
        vendor_names.append(f'Vendor{vendor_index}')

        
    #combined_vendor_data = pd.concat(all_vendor_data)
    combined_vendor_data = pd.concat(all_vendor_data)
    #print("Combined Vendor Data:")
    #print(combined_vendor_data.head())
    #print("Combined Vendor Data shape:", combined_vendor_data.shape)

    # Read the conversion data with NDC and package size
    conversion_columns = ['DRUG NAME', 'ITEM NO', 'NDC #', 'PKG SIZE','PRICE']
    conversion_data = pd.read_excel(conversion_path, usecols=conversion_columns, dtype={'NDC #': str})
    if 'PRICE' in conversion_data.columns:
        conversion_data['PRICE'] = conversion_data['PRICE'].round(0)

    #print("Conversion Data:")
    #print(conversion_data.head())
    #print("Conversion Data shape:", conversion_data.shape)
    
    # Ensure NDC numbers are treated as strings and remove hyphens
    combined_bestrx_data['NDC #'] = combined_bestrx_data['NDC #'].str.replace("-", "").str.zfill(11)
    combined_vendor_data['NDC #'] = combined_vendor_data['NDC #'].str.replace("-", "").str.zfill(11)
    conversion_data['NDC #'] = conversion_data['NDC #'].str.replace("-", "").str.zfill(11)

    

    # Create a mapping for item number and package sizes
    item_no_mapping = conversion_data.set_index('NDC #')['ITEM NO'].to_dict()
    pkg_size_mapping = conversion_data.set_index('NDC #')['PKG SIZE'].to_dict()
    

    # Add package size and item to combined BestRx data
    combined_bestrx_data['Item Number'] = combined_bestrx_data['NDC #'].map(item_no_mapping)
    combined_bestrx_data['Package Size'] = combined_bestrx_data['NDC #'].map(pkg_size_mapping)
    

    # Convert the quantity of tablets to the number of packages
    combined_bestrx_data['Package size'] = combined_bestrx_data['Quantity'] / combined_bestrx_data['Package Size']

    missing_items = combined_bestrx_data[combined_bestrx_data['Item Number'].isnull()][['NDC #', 'Drug Name']].drop_duplicates()


    # Aggregate the number of used bottles for each NDC in combined BestRx data
    bestrx_aggregated = combined_bestrx_data.groupby(['NDC #', 'Drug Name', 'Insurance']).agg({'Package size': 'sum', 'Quantity':'sum', 'Total': 'sum'}).reset_index()
    bestrx_aggregated = bestrx_aggregated.sort_values(by='Drug Name')
    #bestrx_aggregated = combined_bestrx_data.groupby(['NDC', 'Drug Name']).agg({'Package size': 'sum', 'Quantity':'sum'}).reset_index()
    #print("BestRx Aggregated Data:")
    #print(bestrx_aggregated.head())
    #print("Aggregated BestRx Data shape:", bestrx_aggregated.shape)
    
    # Aggregate the number of shipped bottles for each NDC in filtered Kinray data
    #kinray_aggregated = kinray_filtered.groupby('NDC')['Shipped'].sum().reset_index()
    #combined_vendor_data['Shipped'].fillna(0, inplace=True)
    combined_vendor_data['Shipped'] = combined_vendor_data['Shipped'].fillna(0)
    vendor_aggregated = combined_vendor_data.groupby(['NDC #', 'Vendor']).agg({'Shipped': 'sum'}).reset_index()
    #print("Vendor Aggregated Data:")
    #print(vendor_aggregated.head())

    
    vendor_pivot = vendor_aggregated.pivot(index='NDC #', columns='Vendor', values='Shipped').fillna(0).reset_index()
    #print("Vendor Pivot Data:")
    #print(vendor_pivot.head())

    #print("Vendor Aggregated Data:")
    #print(vendor_aggregated.head())
    #print("Vendor Pivot Data:")
    #print(vendor_pivot.head())
    
    # Merge the aggregated BestRx and Kinray data on NDC number
    merged_data = pd.merge(bestrx_aggregated, vendor_pivot, on='NDC #', how='left')
    #print("Merged Data:")
    #print(merged_data.head())

    # Fill NaN values with 0 and ensure numeric type
    for vendor in vendor_names:
        if vendor in merged_data.columns:
            merged_data[vendor] = pd.to_numeric(merged_data[vendor], errors='coerce').fillna(0)
            
    merged_data['Total Purchased'] = merged_data[vendor_names].sum(axis=1)
    

    
    #merged_data['Total Purchased'] = merged_data['Shipped']

    # Pivot to create columns for each insurance company's difference
    pivot_data = merged_data.pivot_table(index=['NDC #', 'Drug Name'], columns='Insurance', values=['Package size', 'Quantity', 'Total'], aggfunc='sum').fillna(0).infer_objects()
    pivot_data.columns = [f'{col[1]}_{col[0][0].upper()}' for col in pivot_data.columns]
    pivot_data = pivot_data.reset_index()
     
    # Print column names after pivot
    #print("Columns after pivot:")
    #print(pivot_data.columns)

     # Merge pivot data with total purchased
    final_data = pd.merge(pivot_data, merged_data[['NDC #', 'Total Purchased'] + vendor_names], on='NDC #', how='left').fillna(0)

    #print("Final Data Before Dropping Duplicates:")
    #print(final_data.head())
    

    # Add item number and package size to the final data
    final_data['Item Number'] = final_data['NDC #'].map(item_no_mapping)
    final_data['Package Size'] = final_data['NDC #'].map(pkg_size_mapping)

    # Calculate differences for each insurance
    for insurance in insurance_paths.keys():
        final_data[f'{insurance}_D'] = final_data['Total Purchased'] - final_data.get(f'{insurance}_P', 0)

    price_mapping = conversion_data.set_index('NDC #')['PRICE'].to_dict()
    final_data['PRICE'] = final_data['NDC #'].map(price_mapping)  # Add price column to the final data
    final_data['Total Order Price'] = abs(final_data['CVS_D']) * final_data['PRICE']  # Calculate the order price

    for insurance in insurance_paths.keys():
        final_data[f'{insurance}_Pur'] = final_data.get(f'{insurance}_P', 0) * final_data['PRICE']

    for insurance in insurance_paths.keys():
        final_data[f'{insurance}_Diff$'] = final_data.get(f'{insurance}_T', 0) - final_data.get(f'{insurance}_Pur', 0)

    
    #print("Columns in the final_data:")
    #print(final_data.columns)
    
    
    #Excel 
    
    
    desired_columns = [
    'Item Number',
    'NDC #', 
    'Drug Name',
    'Package Size'
] + vendor_names + [
    'Total Purchased'
] + \
[f'{insurance}_Q' for insurance in insurance_paths.keys()] + \
[f'{insurance}_P' for insurance in insurance_paths.keys()] + \
[f'{insurance}_D' for insurance in insurance_paths.keys()] + \
[f'{insurance}_T' for insurance in insurance_paths.keys()] + \
[f'{insurance}_Pur' for insurance in insurance_paths.keys()] +\
[f'{insurance}_Diff$' for insurance in insurance_paths.keys()]
    initial_row_count = final_data.shape[0]


    #print(type(final_data))
    #print(final_data[-1])
    #print(final_data)
    final_data = final_data.drop_duplicates(subset=['NDC #', 'Drug Name'])

    # Sort the final data by Drug Name in ascending order
    sorted_data = final_data[desired_columns].sort_values(by='Drug Name')

   #with open('final_data.csv', 'w', newline='', encoding='utf-8') as file:
        #writer = csv.writer(file)
        #writer.writerows(sorted_data)
        #file.write(sorted_data)

    #print("sorted data")    
    #print(sorted_data)
    #print(sorted_data[-1])
    # Check for dropped rows after sorting
    sorted_row_count = sorted_data.shape[0]
    if initial_row_count != sorted_row_count:
        dropped_data.append(('sorted_data', initial_row_count - sorted_row_count))

    
    #Save the sorted data to a new Excel file
    output_file = os.path.join(os.path.expanduser('~'), 'Downloads', f'{pharmacy_name} ({date_range}).xlsx')

    sorted_data.to_excel(output_file, index=False, float_format="%.3f")
    print(f"Processed file saved at: {output_file}")  # Debugging line

    written_data = pd.read_excel(output_file)
    if not os.path.exists(output_file):
        raise FileNotFoundError(f"Processed file not found at {output_file}")


    
    #Set column widths using openpyxl
    wb = load_workbook(output_file)
    ws = wb.active
    

    # Merge the first row and set the pharmacy name and date range in the center
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(desired_columns))
    cell = ws.cell(row=1, column=1)
    cell.value = f"{pharmacy_name} ({date_range})"
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(size=35, bold=True)
    ws.row_dimensions[1].height = 60
    

    
    # Move the data down by one row
    ws.insert_rows(2)
    #Move the data down by one row
    ws.insert_rows(3)
    
    #Explicitly set the headers in the second row
    for col_num, header in enumerate(desired_columns, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(bold=False, size = 15)


    #Dynamically calculate the start and end columns for each merged cell
    def get_column_index(ws, header_name):
        for cell in ws[3]:
            if cell.value == header_name:
                return cell.col_idx
        return None

    #Get the starting column indices for merging
    quantity_billed_start_col = get_column_index(ws, 'ALL_PBM_Q')
    if quantity_billed_start_col is None:
        raise ValueError("Header 'All_Pbm_Q' not found in the worksheet")

    package_size_billed_start_col = get_column_index(ws, 'ALL_PBM_P')
    if package_size_billed_start_col is None:
        raise ValueError("Header 'ALL_PBM_P' not found in the worksheet")

    package_size_difference_start_col = get_column_index(ws, 'ALL_PBM_D')
    if package_size_difference_start_col is None:
        raise ValueError("Header 'ALL_PBM_D' not found in the worksheet")

    quantity_dollar_billed_start_col = get_column_index(ws, 'ALL_PBM_T')
    if quantity_dollar_billed_start_col is None:
        raise ValueError("Header 'ALL_PBM_T' not found in the worksheet")

    
    quantity_dollar_purchased_start_col = get_column_index(ws, 'ALL_PBM_Pur')
    if quantity_dollar_purchased_start_col is None:
        raise ValueError("Header 'ALL_PBM_Pur' not found in the worksheet")

    quantity_dollar_difference_start_col = get_column_index(ws, 'ALL_PBM_Diff$')
    if quantity_dollar_difference_start_col is None:
        raise ValueError("Header 'ALL_PBM_Diff$' not found in the worksheet")
    

    # Calculate the end columns dynamically based on the number of insurance columns
    quantity_billed_end_col = quantity_billed_start_col + len(insurance_paths) - 1
    package_size_billed_end_col = package_size_billed_start_col + len(insurance_paths) - 1
    package_size_difference_end_col = package_size_difference_start_col + len(insurance_paths) - 1
    quantity_dollar_billed_end_col = quantity_dollar_billed_start_col + len(insurance_paths) - 1
    quantity_dollar_purchased_end_col = quantity_dollar_purchased_start_col + len(insurance_paths) - 1
    quantity_dollar_difference_end_col = quantity_dollar_difference_start_col + len(insurance_paths) - 1

    # Merge cells and set the header values dynamically
    ws.merge_cells(start_row=2, start_column=quantity_billed_start_col, end_row=2, end_column=quantity_billed_end_col)
    cell = ws.cell(row=2, column=quantity_billed_start_col)
    cell.value = "Quantity Billed"
    cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(start_row=2, start_column=package_size_billed_start_col, end_row=2, end_column=package_size_billed_end_col)
    cell = ws.cell(row=2, column=package_size_billed_start_col)
    cell.value = "Package size Billed"
    cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(start_row=2, start_column=package_size_difference_start_col, end_row=2, end_column=package_size_difference_end_col)
    cell = ws.cell(row=2, column=package_size_difference_start_col)
    cell.value = "Package size Difference"
    cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(start_row=2, start_column=quantity_dollar_billed_start_col, end_row=2, end_column=quantity_dollar_billed_end_col)
    cell = ws.cell(row=2, column=quantity_dollar_billed_start_col)
    cell.value = "$$ Paid"
    cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(start_row=2, start_column=quantity_dollar_purchased_start_col, end_row=2, end_column=quantity_dollar_purchased_end_col)
    cell = ws.cell(row=2, column=quantity_dollar_purchased_start_col)
    cell.value = "$$ Purchased"
    cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(start_row=2, start_column=quantity_dollar_difference_start_col, end_row=2, end_column=quantity_dollar_difference_end_col)
    cell = ws.cell(row=2, column=quantity_dollar_difference_start_col)
    cell.value = "$$ Difference"
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    
    # Set the desired column widths
    column_widths = {
        'A': 10,  # Item Number
        'B': 15,  # NDC
        'C': 70, #Drug Name
    }

    #Setting up width for the other columns
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    # Set widths for dynamic columns
    for col_num in range(4, len(desired_columns) + 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 7
        

    # Set the height for the first row
    ws.row_dimensions[1].height = 35

    # Set header styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="D0CECE", end_color="D0CECE", fill_type="solid")
    # Set border style
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    thick_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    # Enable text wrapping for the second row
    for cell in ws[3]:
        if cell.col_idx > 3:
            cell.alignment = Alignment(text_rotation=90, horizontal='center', wrap_text=True)
            cell.font = Font(bold=False, size = 14, name='Calibri')
            cell.fill = header_fill
            cell.border = thin_border
        else:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = header_fill
            cell.border = thin_border

    
    ws.row_dimensions[3].height = 100
    # Freeze the first row
    #ws.freeze_panes = 'A4'
    ws.freeze_panes = 'E4' 
    # Center align all data
    for row in ws.iter_rows(min_row=4):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
    # Set the first two columns to left alignment
    for row in ws.iter_rows(min_row=3):
        row[0].alignment = Alignment(horizontal='left')
        row[1].alignment = Alignment(horizontal='left')
        row[2].alignment = Alignment(horizontal='left')

    #Setting up thick border for columns: F to K
    start_col = 6  # Column F
    end_col = 11   # Column K
    start_row = 3
    end_row = ws.max_row
    # Function to apply thick border to all cells
                
    def apply_thick_border(ws, start_col, end_col, start_row, end_row):
       # Apply the thick border to the top row
       for col_num in range(start_col, end_col + 1):
           cell = ws.cell(row=start_row, column=col_num)
           cell.border = Border(
               top=thick_border.top,
               left=cell.border.left,
               right=cell.border.right,
               bottom=cell.border.bottom
           )

       # Apply the thick border to the bottom row
       for col_num in range(start_col, end_col + 1):
           cell = ws.cell(row=end_row, column=col_num)
           cell.border = Border(
               bottom=thick_border.bottom,
               left=cell.border.left,
               right=cell.border.right,
               top=cell.border.top
           )

       # Apply the thick border to the left column
       for row_num in range(start_row, end_row + 1):
           cell = ws.cell(row=row_num, column=start_col)
           cell.border = Border(
               left=thick_border.left,
               top=cell.border.top,
               right=cell.border.right,
               bottom=cell.border.bottom
           )

       # Apply the thick border to the right column
       for row_num in range(start_row, end_row + 1):
           cell = ws.cell(row=row_num, column=end_col)
           cell.border = Border(
               right=thick_border.right,
               top=cell.border.top,
               left=cell.border.left,
               bottom=cell.border.bottom
           )
    start_row = 1
    end_row = ws.max_row

    def style_sheet(ws):
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border

    
    def get_column_indices(ws, column_names):
        indices = []
        for col in ws[3]:
            if col.value in column_names:
                indices.append(col.col_idx)
        return indices
    total_purchased_col = get_column_index(ws, 'Total Purchased')
    if total_purchased_col is None:
        raise ValueError("Header 'Total Purchased' not found in the worksheet")
    
    # Apply thick border for specific column ranges
    quantity_billed_indices = get_column_indices(ws, [f'{insurance}_Q' for insurance in insurance_paths.keys()])
    package_size_billed_indices = get_column_indices(ws, [f'{insurance}_P' for insurance in insurance_paths.keys()])
    package_size_difference_indices = get_column_indices(ws, [f'{insurance}_D' for insurance in insurance_paths.keys()])
    total_purchased_indices = [total_purchased_col]
    dollar_billed_size_difference_indices = get_column_indices(ws, [f'{insurance}_T' for insurance in insurance_paths.keys()])
    dollar_purchased_difference_indices = get_column_indices(ws, [f'{insurance}_Pur' for insurance in insurance_paths.keys()])
    dollar_purchased_difference_indices_ind = get_column_indices(ws, [f'{insurance}_Diff$' for insurance in insurance_paths.keys()])
    
    
    def apply_thick_border_to_groups(ws, column_groups, start_row, end_row):
        for group in column_groups:
            if group:
                start_col = group[0]
                end_col = group[-1]
                apply_thick_border(ws, start_col, end_col, start_row, end_row)

      

    
    for cell in ws[3]:
       cell.border = thin_border
       
    thin_border = Border(left=Side(style='thin', color="A9A9A9"), right=Side(style='thin', color="A9A9A9"), top=Side(style='thin', color="A9A9A9"), bottom=Side(style='thin', color="A9A9A9"))
    # Set up styles
    cell_fill_red = PatternFill(start_color="F88379", end_color="F88379", fill_type="solid")
    row_fill_blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    
    package_size_diff_columns = [get_column_index(ws, f'{insurance}_D') for insurance in insurance_paths.keys()]
    dollar_diff_columns = [get_column_index(ws, f'{insurance}_Diff$') for insurance in insurance_paths.keys()]

    # Highlight rows and cells based on conditions
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        has_negative = False
        for cell in row:
            if cell.col_idx in package_size_diff_columns and isinstance(cell.value, (int, float)) and cell.value < 0:
                cell.fill = cell_fill_red
                has_negative = True
            # Check for negative value in $$ Difference columns
            elif cell.col_idx in dollar_diff_columns and isinstance(cell.value, (int, float)) and cell.value < 0:
                cell.fill = cell_fill_red
                
            cell.border = thin_border
        
        if has_negative:
            for cell in row:
                if cell.fill != cell_fill_red:# Preserve red cells
                    cell.fill = row_fill_blue
                cell.border = thin_border
                
    # Grouping column indices
    column_groups = [quantity_billed_indices, package_size_billed_indices, package_size_difference_indices, total_purchased_indices, dollar_billed_size_difference_indices, dollar_purchased_difference_indices, dollar_purchased_difference_indices_ind]
    apply_thick_border_to_groups(ws, column_groups, start_row, end_row)
    apply_thick_border(ws, start_col=1, end_col=1, start_row=start_row, end_row = end_row)
    apply_thick_border(ws, start_col=2, end_col=2, start_row=start_row, end_row = end_row)
    apply_thick_border(ws, start_col=3, end_col=3, start_row=start_row, end_row = end_row)
    apply_thick_border(ws, start_col=4, end_col=4, start_row=start_row, end_row = end_row)  

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0, right=0, top=0, bottom=0, header=0, footer=0)

    # Set the title of the active worksheet
    ws.title = "Processed Data"
            
    # Inside the process_files function
    # After saving the workbook with the processed data:

    # Add the "Needs to be Ordered" sheet
    add_max_difference_sheet(wb, final_data, insurance_paths)
    min_difference_sheet(wb, final_data, insurance_paths)
    #add_needs_to_order_sheet(wb, final_data, conversion_data) 
    #add_do_not_order(wb, final_data)
    add_missing_items_sheet(wb, missing_items)
    create_never_ordered_check_sheet(wb, final_data)
    

    for sheet in wb.worksheets:
        #adding row height as 20
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):    
            for cell in row:
                if isinstance(cell.value, float):  # Check if the cell contains a float value
                    cell.value = round(cell.value, 2) 
        
    # Set the title in the first row based on the sheet title
        if sheet.title == "Processed Data":
            sheet.cell(row=1, column=1).value = f"{pharmacy_name} ({date_range})"
            cell = sheet.cell(row=1, column=1)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=35, bold=True)
            for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row):
                sheet.row_dimensions[row[0].row].height = 20
            sheet.page_setup.orientation = "landscape"
        elif sheet.title == "Needs to be Ordered":
            sheet.cell(row=1, column=1).value = f"{pharmacy_name} ({date_range}) - NTO CVS"
            cell = sheet.cell(row=1, column=1)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=25, bold=True)
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                sheet.row_dimensions[row[0].row].height = 20
            sheet.page_setup.orientation = "landscape"
        elif sheet.title == "Missing Items":
            sheet.cell(row=1, column=1).value = f"{pharmacy_name} ({date_range}) - Missing items, To be updated in master file"
            cell = sheet.cell(row=1, column=1)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=15, bold=True)
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                sheet.row_dimensions[row[0].row].height = 20
            sheet.page_setup.orientation = "landscape"
        elif sheet.title == "Do Not Order CVS":
            sheet.cell(row=1, column=1).value = f"{pharmacy_name} ({date_range}) - DNO CVS"
            cell = sheet.cell(row=1, column=1)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=25, bold=True)
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                sheet.row_dimensions[row[0].row].height = 20
            sheet.page_setup.orientation = "landscape"

        elif sheet.title == "Needs to be ordered - All":
            sheet.cell(row=1, column=1).value = f"{pharmacy_name} ({date_range}) - Needs to ordered - ALL"
            cell = sheet.cell(row=1, column=1)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=25, bold=True)
            for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row):
                sheet.row_dimensions[row[0].row].height = 20
            sheet.page_setup.orientation = "landscape"
                
        elif sheet.title == "Do Not Order - ALL":#Do Not Order - All
            sheet.cell(row=1, column=1).value = f"{pharmacy_name} ({date_range})-Do Not Order"
            cell = sheet.cell(row=1, column=1)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=25, bold=True)
            for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row):
                sheet.row_dimensions[row[0].row].height = 20
            sheet.page_setup.orientation = "portrait"

        elif sheet.title == "Never Ordered  - Check":
            sheet.cell(row=1, column=1).value = f"{pharmacy_name} ({date_range})-Never Ordered Package - Check"
            cell = sheet.cell(row=1, column=1)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=25, bold=True)
            for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row):
                sheet.row_dimensions[row[0].row].height = 20

            sheet.page_setup.orientation = "landscape"


                
        # Set the first two rows to repeat on each printed page
        sheet.print_title_rows = '1:2'
    
        # Set footer with page numbers
        sheet.oddFooter.left.text = "Page &P of &N"
        sheet.oddFooter.left.size = 8  # Font size for footer
        sheet.oddFooter.left.font = "Arial,Bold"
    
        # Optional: Set page margins and other print options for all sheets
        sheet.page_margins = PageMargins(left=0, right=0, top=0, bottom=0, header=0, footer=0.1)
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.print_options.horizontalCentered = True
        sheet.print_options.verticalCentered = True
        sheet.print_options.gridLines = True
        
        
    start_row = 4  # Assuming data starts from row 4
    end_row = ws.max_row  # Last row of data
    add_autosum(ws, insurance_paths, start_row, end_row)
    
    # Dynamically adjust widths for specific AutoSum columns
    columns_to_adjust = []
    for insurance in insurance_paths.keys():
        t_col = get_column_letter(get_column_index(ws, f'{insurance}_T'))
        pur_col = get_column_letter(get_column_index(ws, f'{insurance}_Pur'))
        diff_col = get_column_letter(get_column_index(ws, f'{insurance}_Diff$'))

        if t_col:
            columns_to_adjust.append(t_col)
        if pur_col:
            columns_to_adjust.append(pur_col)
        if diff_col:
            columns_to_adjust.append(diff_col)

    # Adjust only these specific columns
    adjust_specific_columns(ws, columns_to_adjust)
        
    ws.protection.sheet = True
    wb.save(output_file)

    

    return output_file

if __name__ == '__main__':
    window = webview.create_window('Pharmacy Data Processing Application with price', app, width=800, height=screen_height)
    webview.start()
